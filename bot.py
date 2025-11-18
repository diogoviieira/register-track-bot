import logging
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os
import math
import shutil
import tempfile
from filelock import FileLock

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

# Conversation states
CATEGORY, SUBCATEGORY, AMOUNT, DESCRIPTION, DATE_INPUT, EDIT_FIELD = range(6)

# Excel file path
EXCEL_FILE = "expenses.xlsx"
LOCK_FILE = "expenses.xlsx.lock"

# Excel column indices (0-indexed for reading)
class ExcelColumns:
    DATE = 0
    TIME = 1
    CATEGORY = 2
    SUBCATEGORY = 3
    AMOUNT = 4
    DESCRIPTION = 5
    
    # For writing (1-indexed)
    DATE_WRITE = 1
    TIME_WRITE = 2
    CATEGORY_WRITE = 3
    SUBCATEGORY_WRITE = 4
    AMOUNT_WRITE = 5
    DESCRIPTION_WRITE = 6

# Month mappings (English, Portuguese, and numbers)
MONTH_MAPPINGS = {
    'january': '01', 'janeiro': '01', '1': '01',
    'february': '02', 'fevereiro': '02', '2': '02',
    'march': '03', 'março': '03', 'marco': '03', '3': '03',
    'april': '04', 'abril': '04', '4': '04',
    'may': '05', 'maio': '05', '5': '05',
    'june': '06', 'junho': '06', '6': '06',
    'july': '07', 'julho': '07', '7': '07',
    'august': '08', 'agosto': '08', '8': '08',
    'september': '09', 'setembro': '09', '9': '09',
    'october': '10', 'outubro': '10', '10': '10',
    'november': '11', 'novembro': '11', '11': '11',
    'december': '12', 'dezembro': '12', '12': '12'
}

MONTH_NAMES = {
    '01': 'January', '02': 'February', '03': 'March',
    '04': 'April', '05': 'May', '06': 'June',
    '07': 'July', '08': 'August', '09': 'September',
    '10': 'October', '11': 'November', '12': 'December'
}

# Validation constants
MAX_AMOUNT = 999999

# Common prompts
DATE_FORMAT_PROMPT = (
    "📅 Enter the date {action} (format: DD/MM/YY)\n\n"
    "Example: 15/11/25\n"
    "Or type /cancel to abort."
)

# Main categories
CATEGORIES = [
    ["Home", "Car"],
    ["Lazer", "Travel"],
    ["Needs", "Health"],
    ["Streaming", "Subscriptions"],
    ["Others"]
]

# Subcategories for each main category
SUBCATEGORIES = {
    "Home": [
        ["Rent", "Light"],
        ["Water", "Net"],
        ["Me Mimei", "Other"]
    ],
    "Car": [
        ["Fuel", "Insurance"],
        ["Maintenance", "Parking"],
        ["Via Verde", "Other"]
    ],
    "Lazer": [
        ["Entertainment", "Dining Out"],
        ["Movies/Shows", "Hobbies"],
        ["Coffees", "Other"]
    ],
    "Travel": [
        ["Flights", "Hotels"],
        ["Transportation", "Food"],
        ["Activities", "Other"]
    ],
    "Streaming": [
        ["Prime", "Netflix"],
        ["Disney+", "HBO"]
    ],
    "Subscriptions": [
        ["Patreon", "iCloud"],
        ["Spotify", "F1 TV"],
        ["Telemóvel"]
    ],
    "Needs": [
        ["Groceries", "Clothing"],
        ["Personal Care", "Setup"],
        ["Other"]
    ],
    "Health": [
        ["Doctor", "Pharmacy"],
        ["Hospital", "Gym"],
        ["Supplements", "Other"]
    ],
    "Others": [
        ["Gifts", "Pet"],
        ["Mi Mimei", "Other"]
    ]
}

# Categories/subcategories that don't need description input
# Description will be auto-filled as "Category - Subcategory"
AUTO_DESCRIPTION = {
    "Home": ["Rent", "Light", "Water", "Net"],
    "Car": ["Fuel", "Insurance", "Via Verde"],
    "Streaming": "all",  # All subcategories in Streaming
    "Subscriptions": "all"  # All subcategories in Subscriptions
}


def should_skip_description(category: str, subcategory: str) -> bool:
    """Check if description should be auto-filled for this category/subcategory"""
    if category not in AUTO_DESCRIPTION:
        return False
    
    auto_subs = AUTO_DESCRIPTION[category]
    if auto_subs == "all":
        return True
    
    return subcategory in auto_subs


def get_today_date() -> str:
    """Get today's date in YYYY-MM-DD format"""
    return datetime.now().strftime("%Y-%m-%d")


def get_file_lock(timeout=10):
    """Get a file lock for Excel operations"""
    return FileLock(LOCK_FILE, timeout=timeout)


def format_success_message(category: str, subcategory: str, amount: float, description: str, target_date: str = None) -> str:
    """Format a standardized success message for saved expenses"""
    date_msg = f" for {target_date}" if target_date else ""
    return (
        f"✅ Expense saved successfully{date_msg}!\n\n"
        f"📋 Category: {category}\n"
        f"🏷️ Subcategory: {subcategory}\n"
        f"💵 Amount: €{amount:.2f}\n"
        f"📝 Description: {description}\n\n"
        "Use /add to add another expense or /view to see today's expenses."
    )


async def handle_error(update: Update, error: Exception, operation: str, logger_instance=logger):
    """Centralized error handling for operations"""
    logger_instance.error(f"Error {operation}: {error}")
    await update.message.reply_text(f"Error {operation}. Please try again.")


def format_expense_line(row, values_only=True) -> str:
    """Format a single expense line for display"""
    if values_only:
        # Row is tuple of values
        return f"• {row[ExcelColumns.CATEGORY]} > {row[ExcelColumns.SUBCATEGORY]}: €{row[ExcelColumns.AMOUNT]:.2f} - {row[ExcelColumns.DESCRIPTION]}"
    else:
        # Row is openpyxl Row object with .value attributes
        return f"• {row[ExcelColumns.CATEGORY].value} > {row[ExcelColumns.SUBCATEGORY].value}: €{row[ExcelColumns.AMOUNT].value:.2f} - {row[ExcelColumns.DESCRIPTION].value}"


def format_expense_numbered(index: int, row) -> str:
    """Format a numbered expense line for selection lists"""
    return f"{index}. {row[ExcelColumns.CATEGORY].value} > {row[ExcelColumns.SUBCATEGORY].value}: €{row[ExcelColumns.AMOUNT].value:.2f} - {row[ExcelColumns.DESCRIPTION].value}"


def get_expenses_for_date(target_date: str, lock: FileLock, values_only=True):
    """Load expenses for a specific date with file locking"""
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        if values_only:
            expenses = [
                row for row in ws.iter_rows(min_row=2, values_only=True)
                if row[ExcelColumns.DATE] == target_date
            ]
        else:
            expenses = [
                (idx, row) for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2)
                if row[ExcelColumns.DATE].value == target_date
            ]
        
        return expenses


def safe_save_workbook(wb):
    """Safely save workbook with transaction safety using temp file"""
    try:
        # Save to temp file first
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp:
            tmp_path = tmp.name
        
        # Save workbook to temp file
        wb.save(tmp_path)
        
        # Only replace original if save succeeded
        shutil.move(tmp_path, EXCEL_FILE)
        return True
    except Exception as e:
        # Clean up temp file if it exists
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass
        logger.error(f"Error in safe save: {e}")
        return False


def init_excel():
    """Initialize Excel file if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["Date", "Time", "Category", "Subcategory", "Amount", "Description"])
        wb.save(EXCEL_FILE)
        logger.info(f"Created new Excel file: {EXCEL_FILE}")


def save_expense(category: str, subcategory: str, amount: float, description: str, custom_date: str = None):
    """Save expense to Excel file with optional custom date"""
    lock = get_file_lock()
    try:
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            
            if custom_date:
                date_str = custom_date
            else:
                date_str = datetime.now().strftime("%Y-%m-%d")
            
            time_str = datetime.now().strftime("%H:%M:%S")
            
            ws.append([date_str, time_str, category, subcategory, amount, description])
            
            if safe_save_workbook(wb):
                logger.info(f"Saved expense: {category} > {subcategory} - €{amount} on {date_str}")
                return True
            return False
    except Exception as e:
        logger.error(f"Error saving expense: {e}")
        return False


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show help message with all available commands"""
    await update.message.reply_text(
        "📋 **Available Commands:**\n\n"
        "** Today's Expenses **\n"
        "/add - Add expense for today\n"
        "/view - View today's expenses\n"
        "/summary - Get today's summary\n"
        "/edit - Edit an expense from today\n"
        "/delete - Delete today's expense\n\n"
        "** Specific Date Operations **\n"
        "/add_d - Add expense for a specific date\n"
        "/view_d - View expenses for a specific date\n"
        "/edit_d - Edit expense from a specific date\n"
        "/delete_d - Delete expense from a specific date\n\n"
        "** Monthly Overview **\n"
        "/month <name> - View expenses for a month\n"
        "Example: /month november or /month 11\n\n"
        "** Other **\n"
        "/help - Show this help message\n"
        "/cancel - Cancel current operation"
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start the conversation and ask for expense category"""
    await update.message.reply_text(
        "Hi! I'm your Daddy, i will register your money moves. 📊\n\n"
        "I'll help you not wasting all your money on Putas e Vinho verde.\n\n"
        "Use /help to see all available commands.\n\n"
        "Let's add an expense! Please select a category:",
        reply_markup=ReplyKeyboardMarkup(CATEGORIES, one_time_keyboard=True),
    )
    return CATEGORY


async def add_expense(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start adding a new expense"""
    await update.message.reply_text(
        "Let's add a new expense! 💰\n\n"
        "Please select a category:",
        reply_markup=ReplyKeyboardMarkup(CATEGORIES, one_time_keyboard=True),
    )
    return CATEGORY


async def category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Store category and ask for subcategory"""
    selected_category = update.message.text
    context.user_data["category"] = selected_category
    
    # Get subcategories for the selected category
    if selected_category in SUBCATEGORIES:
        subcats = SUBCATEGORIES[selected_category]
        await update.message.reply_text(
            f"Category: {selected_category}\n\n"
            "Please select a subcategory:",
            reply_markup=ReplyKeyboardMarkup(subcats, one_time_keyboard=True),
        )
        return SUBCATEGORY
    else:
        # If no subcategories defined, skip to amount
        await update.message.reply_text(
            f"Category: {selected_category}\n\n"
            "Now, enter the amount (numbers only):",
            reply_markup=ReplyKeyboardRemove(),
        )
        return AMOUNT


async def subcategory(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Store subcategory and ask for amount"""
    selected_subcategory = update.message.text
    context.user_data["subcategory"] = selected_subcategory
    
    # Check if this category/subcategory should skip description
    category = context.user_data["category"]
    if should_skip_description(category, selected_subcategory):
        context.user_data["skip_description"] = True
    
    await update.message.reply_text(
        f"Subcategory: {selected_subcategory}\n\n"
        "Now, enter the amount (numbers only):",
        reply_markup=ReplyKeyboardRemove(),
    )
    return AMOUNT


async def amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Store amount and ask for description (or auto-save if description not needed)"""
    try:
        amount_value = float(update.message.text)
        
        # Validate amount
        if not math.isfinite(amount_value):
            await update.message.reply_text(
                "❌ Invalid amount. Please enter a valid number:"
            )
            return AMOUNT
        
        if amount_value <= 0:
            await update.message.reply_text(
                "❌ Amount must be positive! Please enter a positive number:"
            )
            return AMOUNT
        
        if amount_value > MAX_AMOUNT:
            await update.message.reply_text(
                "❌ Amount too large! Please enter a reasonable amount:"
            )
            return AMOUNT
        
        context.user_data["amount"] = amount_value
        
        # Check if we should skip description
        if context.user_data.get("skip_description", False):
            # Auto-fill description and save directly
            category = context.user_data["category"]
            subcategory = context.user_data.get("subcategory", "N/A")
            description = f"{category} - {subcategory}"
            target_date = context.user_data.get("target_date")
            
            if save_expense(category, subcategory, amount_value, description, target_date):
                await update.message.reply_text(
                    format_success_message(category, subcategory, amount_value, description, target_date)
                )
            else:
                await update.message.reply_text(
                    "❌ Sorry, there was an error saving your expense. Please try again."
                )
            
            context.user_data.clear()
            return ConversationHandler.END
        else:
            # Ask for description as usual
            await update.message.reply_text(
                f"Amount: €{amount_value:.2f}\n\n"
                "Please provide a brief description:"
            )
            return DESCRIPTION
    except ValueError:
        await update.message.reply_text(
            "Please enter a valid number for the amount:"
        )
        return AMOUNT


async def description(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Store description and save the expense"""
    context.user_data["description"] = update.message.text
    
    category = context.user_data["category"]
    subcategory = context.user_data.get("subcategory", "N/A")
    amount = context.user_data["amount"]
    description = update.message.text
    target_date = context.user_data.get("target_date")
    
    if save_expense(category, subcategory, amount, description, target_date):
        await update.message.reply_text(
            format_success_message(category, subcategory, amount, description, target_date)
        )
    else:
        await update.message.reply_text(
            "❌ Sorry, there was an error saving your expense. Please try again."
        )
    
    context.user_data.clear()
    return ConversationHandler.END


async def view_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """View today's expenses"""
    lock = get_file_lock()
    try:
        today = get_today_date()
        expenses = get_expenses_for_date(today, lock, values_only=True)
        
        if expenses:
            total = sum(float(exp[ExcelColumns.AMOUNT]) for exp in expenses)
            message = f"📊 Today's Expenses ({today}):\n\n"
            for exp in expenses:
                message += format_expense_line(exp) + "\n"
            message += f"\n💰 Total: €{total:.2f}"
        else:
            message = "No expenses recorded for today."
        
        await update.message.reply_text(message)
    except Exception as e:
        await handle_error(update, e, "viewing expenses")
async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Get expense summary by category"""
    lock = get_file_lock()
    try:
        today = get_today_date()
        expenses = get_expenses_for_date(today, lock, values_only=True)
        
        category_totals = {}
        for row in expenses:
            category = row[ExcelColumns.CATEGORY]
            subcategory = row[ExcelColumns.SUBCATEGORY]
            amount = float(row[ExcelColumns.AMOUNT])
            key = f"{category} > {subcategory}"
            category_totals[key] = category_totals.get(key, 0) + amount
        
        if category_totals:
            message = f"📈 Today's Summary ({today}):\n\n"
            for cat, total in sorted(category_totals.items()):
                message += f"• {cat}: €{total:.2f}\n"
            message += f"\n💰 Grand Total: €{sum(category_totals.values()):.2f}"
        else:
            message = "No expenses recorded for today."
        
        await update.message.reply_text(message)
    except Exception as e:
        await handle_error(update, e, "getting summary")
async def view_month_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """View expenses for a specific month"""
    try:
        # Parse the command to get month name or number
        message_text = update.message.text.lower()
        
        # Extract month from command (e.g., "/month november" or "/month 11")
        parts = message_text.split()
        if len(parts) < 2:
            await update.message.reply_text(
                "📅 Please specify a month.\n\n"
                "Examples:\n"
                "/month november\n"
                "/month 11\n"
                "/month novembro"
            )
            return
        
        month_input = parts[1]
        month_num = MONTH_MAPPINGS.get(month_input)
        
        if not month_num:
            await update.message.reply_text(
                "❌ Invalid month. Please use month name or number (1-12).\n\n"
                "Examples: /month november or /month 11"
            )
            return
        
        # Get current year
        current_year = datetime.now().year
        
        # Load expenses
        lock = get_file_lock()
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        
            expenses = []
            total = 0.0
            category_totals = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ExcelColumns.DATE] and row[ExcelColumns.DATE].startswith(f"{current_year}-{month_num}"):
                    expenses.append(row)
                    total += float(row[ExcelColumns.AMOUNT])
                    
                    # Track category totals
                    key = f"{row[ExcelColumns.CATEGORY]} > {row[ExcelColumns.SUBCATEGORY]}"
                    category_totals[key] = category_totals.get(key, 0) + float(row[ExcelColumns.AMOUNT])
        
        if expenses:
            message = f"📊 Expenses for {MONTH_NAMES[month_num]} {current_year}:\n\n"
            message += f"📋 By Category:\n"
            for cat, cat_total in sorted(category_totals.items()):
                message += f"  • {cat}: €{cat_total:.2f}\n"
            
            message += f"\n💰 Total: €{total:.2f}\n"
            message += f"📝 {len(expenses)} expense(s) recorded"
        else:
            await update.message.reply_text(f"No expenses recorded for {month_input.capitalize()} {current_year}.")
            return
        
        await update.message.reply_text(message)
        
    except Exception as e:
        await handle_error(update, e, "viewing month expenses")


async def show_expenses_for_action(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    target_date: str,
    action: str,
    user_data_key: str
):
    """Generic function to show expenses for delete/edit actions"""
    lock = get_file_lock()
    try:
        expenses = get_expenses_for_date(target_date, lock, values_only=False)
        
        if not expenses:
            await update.message.reply_text(f"No expenses to {action} for {target_date}.")
            return
        
        # Build message with appropriate emoji
        emoji = {"delete": "🗑️", "edit": "✏️"}.get(action, "📋")
        message = f"{emoji} Expenses for {target_date}:\n\n"
        
        for i, (row_idx, row) in enumerate(expenses, start=1):
            message += format_expense_numbered(i, row) + "\n"
        
        message += f"\nReply with the number (1-{len(expenses)}) to {action}, or /cancel to abort."
        
        context.user_data[user_data_key] = expenses
        await update.message.reply_text(message)
        
    except Exception as e:
        await handle_error(update, e, f"showing expenses for {action}")
        # Clean up on error
        context.user_data.pop(user_data_key, None)


async def delete_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show today's expenses and allow user to delete one"""
    today = get_today_date()
    await show_expenses_for_action(update, context, today, "delete", "delete_expenses")


async def handle_delete_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the deletion of an expense by number"""
    lock = get_file_lock()
    try:
        choice = int(update.message.text)
        expenses = context.user_data.get("delete_expenses", [])
        
        if not expenses or choice < 1 or choice > len(expenses):
            await update.message.reply_text(
                f"Invalid choice. Please enter a number between 1 and {len(expenses)}, or /cancel."
            )
            return
        
        # Get the row to delete
        row_idx, row = expenses[choice - 1]
        
        # Load workbook and delete the row with file locking
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.delete_rows(row_idx)
            
            if not safe_save_workbook(wb):
                raise Exception("Failed to save workbook")
        
        # Show confirmation
        await update.message.reply_text(
            f"✅ Deleted expense:\n\n"
            f"📋 Category: {row[ExcelColumns.CATEGORY].value}\n"
            f"🏷️ Subcategory: {row[ExcelColumns.SUBCATEGORY].value}\n"
            f"💵 Amount: €{row[ExcelColumns.AMOUNT].value:.2f}\n"
            f"📝 Description: {row[ExcelColumns.DESCRIPTION].value}\n\n"
            "Expense has been removed."
        )
        
    except ValueError:
        await update.message.reply_text(
            "Please enter a valid number, or /cancel to abort."
        )
    except Exception as e:
        await handle_error(update, e, "deleting expense")
    finally:
        # Always clear the delete context
        context.user_data.pop("delete_expenses", None)


# Edit expense functions

async def edit_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show today's expenses and allow user to edit one"""
    today = get_today_date()
    await show_expenses_for_action(update, context, today, "edit", "edit_expenses")


async def edit_expense_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Edit expense for a specific date"""
    await update.message.reply_text(
        DATE_FORMAT_PROMPT.format(action="to edit an expense")
    )
    context.user_data["editing_for_date"] = True
    return DATE_INPUT


async def show_edit_for_date(update: Update, context: ContextTypes.DEFAULT_TYPE, target_date: str):
    """Show expenses for a specific date and allow editing"""
    await show_expenses_for_action(update, context, target_date, "edit", "edit_expenses")


async def handle_edit_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the selection of an expense to edit"""
    try:
        choice = int(update.message.text)
        expenses = context.user_data.get("edit_expenses", [])
        
        if not expenses or choice < 1 or choice > len(expenses):
            await update.message.reply_text(
                f"Invalid choice. Please enter a number between 1 and {len(expenses)}, or /cancel."
            )
            return
        
        # Store the selected expense for editing
        row_idx, row = expenses[choice - 1]
        context.user_data["edit_row_idx"] = row_idx
        context.user_data["edit_expense_data"] = {
            "category": row[ExcelColumns.CATEGORY].value,
            "subcategory": row[ExcelColumns.SUBCATEGORY].value,
            "amount": row[ExcelColumns.AMOUNT].value,
            "description": row[ExcelColumns.DESCRIPTION].value
        }
        context.user_data.pop("edit_expenses", None)
        
        # Show what can be edited
        await update.message.reply_text(
            f"✏️ Editing expense:\n\n"
            f"📋 Category: {row[ExcelColumns.CATEGORY].value}\n"
            f"🏷️ Subcategory: {row[ExcelColumns.SUBCATEGORY].value}\n"
            f"💵 Amount: €{row[ExcelColumns.AMOUNT].value:.2f}\n"
            f"📝 Description: {row[ExcelColumns.DESCRIPTION].value}\n\n"
            "What would you like to edit?\n"
            "Reply with:\n"
            "• 'amount' - Change the amount\n"
            "• 'description' - Change the description\n"
            "• /cancel - Cancel editing"
        )
        
    except ValueError:
        await update.message.reply_text(
            "Please enter a valid number, or /cancel to abort."
        )
    except Exception as e:
        await handle_error(update, e, "selecting expense for edit")
        # Clean up on error
        context.user_data.pop("edit_expenses", None)
        context.user_data.pop("edit_row_idx", None)
        context.user_data.pop("edit_expense_data", None)


async def handle_edit_field_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the choice of what field to edit"""
    try:
        choice = update.message.text.lower().strip()
        
        if choice == "amount":
            context.user_data["editing_field"] = "amount"
            current_amount = context.user_data["edit_expense_data"]["amount"]
            await update.message.reply_text(
                f"Current amount: €{current_amount:.2f}\n\n"
                "Enter the new amount (numbers only):"
            )
        elif choice == "description":
            context.user_data["editing_field"] = "description"
            current_desc = context.user_data["edit_expense_data"]["description"]
            await update.message.reply_text(
                f"Current description: {current_desc}\n\n"
                "Enter the new description:"
            )
        else:
            await update.message.reply_text(
                "Please reply with 'amount' or 'description', or /cancel to abort."
            )
    except Exception as e:
        await handle_error(update, e, "handling edit field choice")
        # Clean up on error
        context.user_data.pop("edit_row_idx", None)
        context.user_data.pop("edit_expense_data", None)
        context.user_data.pop("editing_field", None)


async def handle_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the new value for the edited field"""
    lock = get_file_lock()
    try:
        field = context.user_data.get("editing_field")
        new_value = update.message.text
        
        if field == "amount":
            new_value = float(new_value)
            
            # Validate amount
            if not math.isfinite(new_value):
                await update.message.reply_text(
                    "❌ Invalid amount. Please enter a valid number:"
                )
                return
            
            if new_value <= 0:
                await update.message.reply_text(
                    "❌ Amount must be positive! Please enter a positive number:"
                )
                return
            
            if new_value > MAX_AMOUNT:
                await update.message.reply_text(
                    "❌ Amount too large! Please enter a reasonable amount:"
                )
                return
        
        # Load workbook and update the expense with file locking
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            row_idx = context.user_data["edit_row_idx"]
            
            if field == "amount":
                ws.cell(row=row_idx, column=ExcelColumns.AMOUNT_WRITE).value = new_value
                context.user_data["edit_expense_data"]["amount"] = new_value
            elif field == "description":
                ws.cell(row=row_idx, column=ExcelColumns.DESCRIPTION_WRITE).value = new_value
                context.user_data["edit_expense_data"]["description"] = new_value
            
            if not safe_save_workbook(wb):
                raise Exception("Failed to save workbook")
        
        # Show confirmation
        data = context.user_data["edit_expense_data"]
        await update.message.reply_text(
            f"✅ Expense updated successfully!\n\n"
            f"📋 Category: {data['category']}\n"
            f"🏷️ Subcategory: {data['subcategory']}\n"
            f"💵 Amount: €{data['amount']:.2f}\n"
            f"📝 Description: {data['description']}"
        )
        
    except ValueError:
        await update.message.reply_text(
            "Please enter a valid number for the amount, or /cancel to abort."
        )
    except Exception as e:
        await handle_error(update, e, "updating expense")
    finally:
        # Always clear edit context
        context.user_data.pop("edit_row_idx", None)
        context.user_data.pop("edit_expense_data", None)
        context.user_data.pop("editing_field", None)


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel the conversation"""
    await update.message.reply_text(
        "Operation cancelled. Use /add to add a new expense.",
        reply_markup=ReplyKeyboardRemove(),
    )
    context.user_data.clear()
    return ConversationHandler.END


# Date-specific functions

async def add_expense_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start adding an expense for a specific date"""
    await update.message.reply_text(
        DATE_FORMAT_PROMPT.format(action="for the expense")
    )
    context.user_data["adding_for_date"] = True
    return DATE_INPUT


async def view_expenses_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """View expenses for a specific date"""
    await update.message.reply_text(
        DATE_FORMAT_PROMPT.format(action="to view expenses")
    )
    context.user_data["viewing_date"] = True
    return DATE_INPUT


async def delete_expense_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Delete expense for a specific date"""
    await update.message.reply_text(
        DATE_FORMAT_PROMPT.format(action="to delete an expense")
    )
    context.user_data["deleting_for_date"] = True
    return DATE_INPUT


async def handle_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle date input and route to appropriate action"""
    date_text = update.message.text.strip()
    
    # Validate date format (DD/MM/YY)
    try:
        parsed_date = datetime.strptime(date_text, "%d/%m/%y")
        date_str = parsed_date.strftime("%Y-%m-%d")  # Store internally as YYYY-MM-DD
    except ValueError:
        await update.message.reply_text(
            "❌ Invalid date format. Please use DD/MM/YY format.\n"
            "Example: 15/11/25\n\n"
            "Or type /cancel to abort."
        )
        return DATE_INPUT
    
    # Store the date
    context.user_data["target_date"] = date_str
    
    # Route based on what action was requested
    if context.user_data.get("adding_for_date"):
        context.user_data.pop("adding_for_date")
        await update.message.reply_text(
            f"📅 Adding expense for {date_str}\n\n"
            "Please select a category:",
            reply_markup=ReplyKeyboardMarkup(CATEGORIES, one_time_keyboard=True),
        )
        return CATEGORY
    
    elif context.user_data.get("viewing_date"):
        context.user_data.pop("viewing_date")
        await view_expenses_for_date(update, context, date_str)
        return ConversationHandler.END
    
    elif context.user_data.get("deleting_for_date"):
        context.user_data.pop("deleting_for_date")
        await show_delete_for_date(update, context, date_str)
        return ConversationHandler.END
    
    elif context.user_data.get("editing_for_date"):
        context.user_data.pop("editing_for_date")
        await show_edit_for_date(update, context, date_str)
        return ConversationHandler.END
    
    return ConversationHandler.END


async def view_expenses_for_date(update: Update, context: ContextTypes.DEFAULT_TYPE, target_date: str):
    """View expenses for a specific date"""
    lock = get_file_lock()
    try:
        expenses = get_expenses_for_date(target_date, lock, values_only=True)
        
        if expenses:
            total = sum(float(exp[ExcelColumns.AMOUNT]) for exp in expenses)
            message = f"📊 Expenses for {target_date}:\n\n"
            for exp in expenses:
                message += format_expense_line(exp) + "\n"
            message += f"\n💰 Total: €{total:.2f}"
        else:
            message = f"No expenses recorded for {target_date}."
        
        await update.message.reply_text(message)
    except Exception as e:
        await handle_error(update, e, "viewing expenses for date")


async def show_delete_for_date(update: Update, context: ContextTypes.DEFAULT_TYPE, target_date: str):
    """Show expenses for a specific date and allow deletion"""
    await show_expenses_for_action(update, context, target_date, "delete", "delete_expenses")


def main():
    """Start the bot"""
    # Initialize Excel file
    init_excel()
    
    # Read bot token from environment variable or config
    BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
    
    if not BOT_TOKEN:
        print("ERROR: Please set TELEGRAM_BOT_TOKEN environment variable")
        print("\nSteps to get your bot token:")
        print("1. Open Telegram and search for @BotFather")
        print("2. Send /newbot command")
        print("3. Follow the instructions to create your bot")
        print("4. Copy the token and set it as environment variable:")
        print("   Windows: setx TELEGRAM_BOT_TOKEN \"your-token-here\"")
        print("   Linux/Mac: export TELEGRAM_BOT_TOKEN=\"your-token-here\"")
        return
    
    # Create application
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Add conversation handler for adding expenses (today or specific date)
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CommandHandler("add", add_expense),
            CommandHandler("add_d", add_expense_date),
        ],
        states={
            DATE_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date_input)],
            CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, category)],
            SUBCATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, subcategory)],
            AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, amount)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Conversation handler for viewing specific date
    view_date_handler = ConversationHandler(
        entry_points=[CommandHandler("view_d", view_expenses_date)],
        states={
            DATE_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date_input)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Conversation handler for deleting from specific date
    delete_date_handler = ConversationHandler(
        entry_points=[CommandHandler("delete_d", delete_expense_date)],
        states={
            DATE_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date_input)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Conversation handler for editing from specific date
    edit_date_handler = ConversationHandler(
        entry_points=[CommandHandler("edit_d", edit_expense_date)],
        states={
            DATE_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date_input)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    application.add_handler(conv_handler)
    application.add_handler(view_date_handler)
    application.add_handler(delete_date_handler)
    application.add_handler(edit_date_handler)
    application.add_handler(CommandHandler("view", view_expenses))
    application.add_handler(CommandHandler("month", view_month_expenses))
    application.add_handler(CommandHandler("summary", summary))
    application.add_handler(CommandHandler("edit", edit_expense))
    application.add_handler(CommandHandler("delete", delete_expense))
    
    # Combined handler for all non-conversation text input
    async def handle_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip()
        
        # Priority 1: Editing field value
        if context.user_data.get("editing_field"):
            await handle_edit_value(update, context)
            return
        
        # Priority 2: Selecting what field to edit
        if context.user_data.get("edit_expense_data") and text.lower() in ["amount", "description"]:
            await handle_edit_field_choice(update, context)
            return
        
        # Priority 3: Selecting expense number to edit
        if context.user_data.get("edit_expenses") and text.isdigit():
            await handle_edit_number(update, context)
            return
        
        # Priority 4: Selecting expense number to delete
        if context.user_data.get("delete_expenses") and text.isdigit():
            await handle_delete_number(update, context)
            return
    
    application.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        handle_text_input
    ))
    
    # Handle unknown commands
    async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(
            "Burro, you can use /help to get a list of all the available commands !"
        )
    
    # Help command handler
    application.add_handler(CommandHandler("help", help_command))
    
    # Unknown command handler - must be last
    application.add_handler(MessageHandler(filters.COMMAND, unknown_command))
    
    # Start the bot
    print("Bot is running... Press Ctrl+C to stop.")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()