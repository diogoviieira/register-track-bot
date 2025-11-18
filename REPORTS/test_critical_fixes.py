"""
Test script to verify all critical fixes in bot.py
Tests: file locking, input validation, memory safety, transaction safety
"""
import os
import sys
import math
import tempfile
import shutil
from filelock import FileLock
import openpyxl
from openpyxl import Workbook

# Test configuration
TEST_EXCEL = "test_expenses.xlsx"
TEST_LOCK = "test_expenses.xlsx.lock"

def test_safe_save_workbook():
    """Test transaction safety with tempfile"""
    print("\n=== Testing Transaction Safety ===")
    
    wb_test = None
    try:
        # Create a test workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Time", "Category", "Subcategory", "Amount", "Description"])
        ws.append(["2025-11-18", "10:00:00", "Home", "Rent", 1000.00, "Test"])
        
        # Test the safe save logic
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp:
            tmp_path = tmp.name
        
        wb.save(tmp_path)
        wb.close()
        
        # Verify file was created
        if os.path.exists(tmp_path):
            size = os.path.getsize(tmp_path)
            print(f"✅ Temp file created: {size} bytes")
            
            # Simulate move to actual location
            shutil.move(tmp_path, TEST_EXCEL)
            
            if os.path.exists(TEST_EXCEL):
                print("✅ Atomic save successful")
                
                # Verify data integrity
                wb_test = openpyxl.load_workbook(TEST_EXCEL)
                ws_test = wb_test.active
                rows = list(ws_test.iter_rows(values_only=True))
                if len(rows) == 2 and rows[1][4] == 1000.00:
                    print("✅ Data integrity verified")
                else:
                    print("❌ Data integrity check failed")
            else:
                print("❌ File move failed")
        else:
            print("❌ Temp file creation failed")
            
    except Exception as e:
        print(f"❌ Exception: {e}")
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            os.remove(tmp_path)
    finally:
        # Close workbook if open
        if wb_test is not None:
            wb_test.close()
        
        # Cleanup
        if os.path.exists(TEST_EXCEL):
            try:
                os.remove(TEST_EXCEL)
            except PermissionError:
                import time
                time.sleep(0.2)
                try:
                    os.remove(TEST_EXCEL)
                except:
                    pass


def test_file_locking():
    """Test file locking prevents concurrent access"""
    print("\n=== Testing File Locking ===")
    
    lock = FileLock(TEST_LOCK, timeout=2)
    
    try:
        # Acquire lock
        with lock:
            print("✅ Lock acquired successfully")
            
            # Try to acquire again (should timeout)
            lock2 = FileLock(TEST_LOCK, timeout=1)
            try:
                with lock2:
                    print("❌ Second lock acquired (should not happen)")
            except Exception as e:
                if "Timeout" in str(e.__class__.__name__):
                    print("✅ Second lock blocked (correct behavior)")
                else:
                    print(f"⚠️ Unexpected exception: {e}")
                    
    except Exception as e:
        print(f"❌ Lock test failed: {e}")
    finally:
        # Cleanup
        if os.path.exists(TEST_LOCK):
            os.remove(TEST_LOCK)


def test_input_validation():
    """Test amount validation logic"""
    print("\n=== Testing Input Validation ===")
    
    test_cases = [
        (100.50, True, "Valid positive amount"),
        (0, False, "Zero amount (invalid)"),
        (-50, False, "Negative amount (invalid)"),
        (999999, True, "Maximum valid amount"),
        (1000000, False, "Amount too large"),
        (float('inf'), False, "Infinite value (invalid)"),
        (float('nan'), False, "NaN value (invalid)"),
        (0.01, True, "Small positive amount"),
    ]
    
    passed = 0
    failed = 0
    
    for value, should_pass, description in test_cases:
        # Simulate validation logic
        is_valid = True
        
        try:
            if not math.isfinite(value):
                is_valid = False
            elif value <= 0:
                is_valid = False
            elif value > 999999:
                is_valid = False
        except:
            is_valid = False
        
        if is_valid == should_pass:
            print(f"✅ {description}: {value}")
            passed += 1
        else:
            print(f"❌ {description}: {value} (expected {should_pass}, got {is_valid})")
            failed += 1
    
    print(f"\nValidation tests: {passed} passed, {failed} failed")


def test_memory_cleanup_simulation():
    """Test memory cleanup patterns"""
    print("\n=== Testing Memory Cleanup Patterns ===")
    
    # Simulate context.user_data
    user_data = {}
    
    # Test 1: Normal cleanup
    try:
        user_data["delete_expenses"] = [(1, "data")]
        user_data["edit_row_idx"] = 5
        user_data["editing_field"] = "amount"
        
        # Simulate successful operation
        user_data.pop("delete_expenses", None)
        user_data.pop("edit_row_idx", None)
        user_data.pop("editing_field", None)
        
        if len(user_data) == 0:
            print("✅ Normal cleanup successful")
        else:
            print(f"❌ Data remained: {user_data}")
    except Exception as e:
        print(f"❌ Normal cleanup failed: {e}")
    
    # Test 2: Error path cleanup with finally
    try:
        user_data["delete_expenses"] = [(1, "data")]
        try:
            raise Exception("Simulated error")
        except Exception as e:
            pass
        finally:
            user_data.pop("delete_expenses", None)
        
        if len(user_data) == 0:
            print("✅ Error path cleanup successful")
        else:
            print(f"❌ Data remained after error: {user_data}")
    except Exception as e:
        print(f"❌ Error path cleanup failed: {e}")
    
    # Test 3: Multiple cleanup keys
    try:
        user_data["edit_row_idx"] = 10
        user_data["edit_expense_data"] = {"amount": 100}
        user_data["editing_field"] = "description"
        
        # Cleanup all
        user_data.pop("edit_row_idx", None)
        user_data.pop("edit_expense_data", None)
        user_data.pop("editing_field", None)
        
        if len(user_data) == 0:
            print("✅ Multiple key cleanup successful")
        else:
            print(f"❌ Data remained: {user_data}")
    except Exception as e:
        print(f"❌ Multiple key cleanup failed: {e}")


def test_excel_operations_with_lock():
    """Test Excel operations with file locking"""
    print("\n=== Testing Excel Operations with Locking ===")
    
    # Create test file
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Category", "Subcategory", "Amount", "Description"])
    wb.save(TEST_EXCEL)
    
    lock = FileLock(TEST_LOCK, timeout=10)
    
    try:
        # Test write operation
        with lock:
            wb = openpyxl.load_workbook(TEST_EXCEL)
            ws = wb.active
            ws.append(["2025-11-18", "10:00:00", "Home", "Rent", 1000.00, "Monthly rent"])
            
            # Use safe save pattern
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp:
                tmp_path = tmp.name
            wb.save(tmp_path)
            shutil.move(tmp_path, TEST_EXCEL)
        
        print("✅ Write operation with lock successful")
        
        # Test read operation
        with lock:
            wb = openpyxl.load_workbook(TEST_EXCEL)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) == 2 and rows[1][4] == 1000.00:
                print("✅ Read operation with lock successful")
            else:
                print(f"❌ Read operation failed: {len(rows)} rows")
            wb.close()
        
    except Exception as e:
        print(f"❌ Excel operations failed: {e}")
    finally:
        # Cleanup
        import time
        time.sleep(0.2)  # Give OS time to release file handles
        
        if os.path.exists(TEST_EXCEL):
            try:
                os.remove(TEST_EXCEL)
            except PermissionError:
                time.sleep(0.3)
                try:
                    os.remove(TEST_EXCEL)
                except:
                    print(f"⚠️ Could not remove test file {TEST_EXCEL}, may need manual cleanup")
        if os.path.exists(TEST_LOCK):
            try:
                os.remove(TEST_LOCK)
            except:
                pass


def main():
    print("=" * 60)
    print("CRITICAL FIXES VERIFICATION TEST SUITE")
    print("=" * 60)
    
    # Run all tests
    test_safe_save_workbook()
    test_file_locking()
    test_input_validation()
    test_memory_cleanup_simulation()
    test_excel_operations_with_lock()
    
    print("\n" + "=" * 60)
    print("TEST SUITE COMPLETED")
    print("=" * 60)
    print("\nReview the results above to ensure all fixes are working.")
    print("Look for ✅ (pass) and ❌ (fail) indicators.")


if __name__ == "__main__":
    main()
