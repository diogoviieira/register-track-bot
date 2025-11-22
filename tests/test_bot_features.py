"""
Comprehensive test script for register-track-bot
Tests all features including expenses, incomes, and helper functions
"""

import sys
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook

# Add bot directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Import bot functions
from bot import (
    # Constants
    ExcelColumns,
    EXCEL_FILE,
    INCOME_FILE,
    MAX_AMOUNT,
    MONTH_MAPPINGS,
    MONTH_NAMES,
    CATEGORIES,
    SUBCATEGORIES,
    AUTO_DESCRIPTION,
    
    # Helper functions
    should_skip_description,
    get_today_date,
    get_file_lock,
    format_expense_line,
    format_expense_numbered,
    format_success_message,
    get_expenses_for_date,
    safe_save_workbook,
    
    # Core functions
    init_excel,
    init_income_excel,
    save_expense,
)

class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    END = '\033[0m'
    BOLD = '\033[1m'

def print_test(name, status, details=""):
    symbol = f"{Colors.GREEN}✓{Colors.END}" if status else f"{Colors.RED}✗{Colors.END}"
    print(f"{symbol} {name}")
    if details:
        print(f"  {Colors.CYAN}{details}{Colors.END}")

def print_section(title):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{title}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}\n")

def test_constants():
    """Test all constants are properly defined"""
    print_section("Testing Constants")
    
    # Test ExcelColumns
    try:
        assert ExcelColumns.DATE == 0
        assert ExcelColumns.TIME == 1
        assert ExcelColumns.CATEGORY == 2
        assert ExcelColumns.SUBCATEGORY == 3
        assert ExcelColumns.AMOUNT == 4
        assert ExcelColumns.DESCRIPTION == 5
        assert ExcelColumns.DATE_WRITE == 1
        assert ExcelColumns.AMOUNT_WRITE == 5
        print_test("ExcelColumns constants", True, "All column indices correct")
    except AssertionError as e:
        print_test("ExcelColumns constants", False, str(e))
    
    # Test MAX_AMOUNT
    try:
        assert MAX_AMOUNT == 999999
        print_test("MAX_AMOUNT constant", True, f"Value: {MAX_AMOUNT}")
    except AssertionError:
        print_test("MAX_AMOUNT constant", False, f"Expected 999999, got {MAX_AMOUNT}")
    
    # Test MONTH_MAPPINGS
    try:
        assert 'january' in MONTH_MAPPINGS
        assert 'janeiro' in MONTH_MAPPINGS
        assert MONTH_MAPPINGS['november'] == '11'
        assert MONTH_MAPPINGS['11'] == '11'
        print_test("MONTH_MAPPINGS", True, f"{len(MONTH_MAPPINGS)} month mappings defined")
    except AssertionError as e:
        print_test("MONTH_MAPPINGS", False, str(e))
    
    # Test MONTH_NAMES
    try:
        assert MONTH_NAMES['01'] == 'January'
        assert MONTH_NAMES['11'] == 'November'
        print_test("MONTH_NAMES", True, f"{len(MONTH_NAMES)} month names defined")
    except AssertionError as e:
        print_test("MONTH_NAMES", False, str(e))
    
    # Test Incomes category exists
    try:
        found_incomes = any('Incomes' in row for row in CATEGORIES)
        assert found_incomes
        print_test("Incomes category in CATEGORIES", True, "Found in main categories")
    except AssertionError:
        print_test("Incomes category in CATEGORIES", False, "Not found")
    
    # Test Incomes subcategories
    try:
        assert 'Incomes' in SUBCATEGORIES
        incomes_subs = SUBCATEGORIES['Incomes']
        flat_subs = [item for sublist in incomes_subs for item in sublist]
        assert 'Refeição' in flat_subs
        assert 'Subsídio' in flat_subs
        assert 'Bónus' in flat_subs
        assert 'Salary' in flat_subs
        assert 'Others' in flat_subs
        print_test("Incomes subcategories", True, f"5 subcategories: {', '.join(flat_subs)}")
    except (AssertionError, KeyError) as e:
        print_test("Incomes subcategories", False, str(e))
    
    # Test AUTO_DESCRIPTION includes Incomes
    try:
        assert 'Incomes' in AUTO_DESCRIPTION
        income_auto = AUTO_DESCRIPTION['Incomes']
        assert 'Refeição' in income_auto
        assert 'Salary' in income_auto
        assert 'Others' not in income_auto  # Others should require description
        print_test("AUTO_DESCRIPTION for Incomes", True, "4 auto-desc, 'Others' requires input")
    except (AssertionError, KeyError) as e:
        print_test("AUTO_DESCRIPTION for Incomes", False, str(e))

def test_helper_functions():
    """Test helper functions"""
    print_section("Testing Helper Functions")
    
    # Test should_skip_description
    try:
        assert should_skip_description("Incomes", "Salary") == True
        assert should_skip_description("Incomes", "Refeição") == True
        assert should_skip_description("Incomes", "Others") == False
        assert should_skip_description("Home", "Rent") == True
        assert should_skip_description("Home", "Other") == False
        print_test("should_skip_description()", True, "Correctly identifies auto-description categories")
    except AssertionError as e:
        print_test("should_skip_description()", False, str(e))
    
    # Test get_today_date
    try:
        today = get_today_date()
        assert isinstance(today, str)
        assert len(today) == 10  # YYYY-MM-DD format
        assert today.count('-') == 2
        datetime.strptime(today, "%Y-%m-%d")  # Validate format
        print_test("get_today_date()", True, f"Returns: {today}")
    except Exception as e:
        print_test("get_today_date()", False, str(e))
    
    # Test format_success_message
    try:
        msg = format_success_message("Test", "SubTest", 100.50, "Test Description")
        assert "Test" in msg
        assert "SubTest" in msg
        assert "100.50" in msg
        assert "Test Description" in msg
        assert "✅" in msg
        print_test("format_success_message()", True, "Formats expense confirmation correctly")
    except AssertionError as e:
        print_test("format_success_message()", False, str(e))
    
    # Test format_expense_line
    try:
        test_row = ("2025-11-22", "10:00:00", "Home", "Rent", 500.00, "Monthly rent")
        line = format_expense_line(test_row, values_only=True)
        assert "Home" in line
        assert "Rent" in line
        assert "500.00" in line
        assert "Monthly rent" in line
        print_test("format_expense_line()", True, f"Output: {line[:50]}...")
    except Exception as e:
        print_test("format_expense_line()", False, str(e))

def test_file_initialization():
    """Test Excel file initialization"""
    print_section("Testing File Initialization")
    
    # Backup existing files if they exist
    backups = []
    for file in [EXCEL_FILE, INCOME_FILE]:
        if os.path.exists(file):
            backup = f"{file}.backup"
            if os.path.exists(backup):
                os.remove(backup)
            os.rename(file, backup)
            backups.append((file, backup))
    
    try:
        # Test expense file initialization
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        init_excel()
        assert os.path.exists(EXCEL_FILE)
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()  # Close workbook before cleanup
        assert headers == ["Date", "Time", "Category", "Subcategory", "Amount", "Description"]
        print_test("init_excel()", True, "Creates expenses.xlsx with correct headers")
    except Exception as e:
        print_test("init_excel()", False, str(e))
    
    try:
        # Test income file initialization
        if os.path.exists(INCOME_FILE):
            os.remove(INCOME_FILE)
        init_income_excel()
        assert os.path.exists(INCOME_FILE)
        wb = openpyxl.load_workbook(INCOME_FILE)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()  # Close workbook before cleanup
        assert headers == ["Date", "Time", "Category", "Subcategory", "Amount", "Description"]
        print_test("init_income_excel()", True, "Creates incomes.xlsx with correct headers")
    except Exception as e:
        print_test("init_income_excel()", False, str(e))
    
    # Restore backups (with proper cleanup)
    import time
    time.sleep(0.1)  # Brief pause to ensure files are released
    for original, backup in backups:
        try:
            if os.path.exists(original):
                os.remove(original)
            if os.path.exists(backup):
                os.rename(backup, original)
        except PermissionError:
            pass  # File still in use, skip restore

def test_save_functions():
    """Test save expense/income functions"""
    print_section("Testing Save Functions")
    
    # Initialize files
    init_excel()
    init_income_excel()
    
    # Test saving expense
    try:
        result = save_expense("Home", "Rent", 500.00, "Test expense", "2025-11-22")
        assert result == True
        
        # Verify it was saved to expenses.xlsx
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        last_row = list(ws.iter_rows(min_row=2, values_only=True))[-1]
        wb.close()
        assert last_row[ExcelColumns.CATEGORY] == "Home"
        assert last_row[ExcelColumns.SUBCATEGORY] == "Rent"
        assert last_row[ExcelColumns.AMOUNT] == 500.00
        print_test("save_expense() for expenses", True, "Saves to expenses.xlsx correctly")
    except Exception as e:
        print_test("save_expense() for expenses", False, str(e))
    
    # Test saving income
    try:
        result = save_expense("Incomes", "Salary", 2000.00, "Incomes - Salary", "2025-11-22")
        assert result == True
        
        # Verify it was saved to incomes.xlsx
        wb = openpyxl.load_workbook(INCOME_FILE)
        ws = wb.active
        last_row = list(ws.iter_rows(min_row=2, values_only=True))[-1]
        wb.close()
        assert last_row[ExcelColumns.CATEGORY] == "Incomes"
        assert last_row[ExcelColumns.SUBCATEGORY] == "Salary"
        assert last_row[ExcelColumns.AMOUNT] == 2000.00
        print_test("save_expense() for incomes", True, "Saves to incomes.xlsx correctly")
    except Exception as e:
        print_test("save_expense() for incomes", False, str(e))
    
    # Test that incomes are NOT in expenses file
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        income_in_expenses = any(
            row[ExcelColumns.CATEGORY] == "Incomes" 
            for row in ws.iter_rows(min_row=2, values_only=True) 
            if row[ExcelColumns.CATEGORY]
        )
        wb.close()
        assert income_in_expenses == False
        print_test("Income/Expense separation", True, "Incomes NOT saved to expenses.xlsx")
    except Exception as e:
        print_test("Income/Expense separation", False, str(e))

def test_data_retrieval():
    """Test data retrieval functions"""
    print_section("Testing Data Retrieval")
    
    # Test get_expenses_for_date
    try:
        lock = get_file_lock()
        expenses = get_expenses_for_date("2025-11-22", lock, values_only=True)
        assert isinstance(expenses, list)
        if len(expenses) > 0:
            assert len(expenses[0]) == 6  # 6 columns
            print_test("get_expenses_for_date()", True, f"Found {len(expenses)} expense(s)")
        else:
            print_test("get_expenses_for_date()", True, "Returns empty list (no expenses)")
    except Exception as e:
        print_test("get_expenses_for_date()", False, str(e))

def run_all_tests():
    """Run all test suites"""
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*60}")
    print(f"  REGISTER-TRACK-BOT COMPREHENSIVE TEST SUITE")
    print(f"  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}{Colors.END}\n")
    
    test_constants()
    test_helper_functions()
    test_file_initialization()
    test_save_functions()
    test_data_retrieval()
    
    print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*60}")
    print(f"  TEST SUITE COMPLETE")
    print(f"{'='*60}{Colors.END}\n")
    
    print(f"{Colors.YELLOW}Note: Bot features requiring Telegram interaction (commands, messages)")
    print(f"      must be tested manually by running the bot and using Telegram.{Colors.END}\n")
    
    print(f"{Colors.CYAN}Features tested:{Colors.END}")
    print(f"  ✓ Constants (ExcelColumns, MAX_AMOUNT, months, categories)")
    print(f"  ✓ Incomes category and subcategories")
    print(f"  ✓ Helper functions (date, formatting, validation)")
    print(f"  ✓ File initialization (expenses.xlsx, incomes.xlsx)")
    print(f"  ✓ Save functions with proper routing")
    print(f"  ✓ Income/Expense separation")
    print(f"  ✓ Data retrieval functions\n")
    
    print(f"{Colors.CYAN}Manual testing required:{Colors.END}")
    print(f"  • /add command with Incomes category")
    print(f"  • /income <month> command")
    print(f"  • /view, /summary, /month (should exclude incomes)")
    print(f"  • All existing expense commands")
    print(f"  • Edit and delete operations\n")

if __name__ == "__main__":
    run_all_tests()
